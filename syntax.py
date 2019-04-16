import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# The following file can help you learn the more common syntax used with openpyxl. Many of the commands will have a print
# function added to them. This is not always needed it is simply so you can run them easily and see the results printed.


############################################# Opening WB and sheets #############################################


# Create a workbook object
wb = openpyxl.load_workbook('example.xlsx')

# Find out the sheet names
print(wb.sheetnames)

# Find active sheet
(print(wb.active))

# Create a sheet object
sheet = wb['Sheet1']


############################################# Reading from Excel #############################################


# Access the value of a cell on a sheet using "A1" format.  This value being column "A" and row "1".
print(sheet["A2"].value)

# Access the value of a specific cell on sheet using numbers for both rows and columns.
print(sheet.cell(row=2, column=2).value)

# Use the above syntax in a loop
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

# Convert a column location from a letter value to number value or visa versa. Needs to be imported.
print(get_column_letter(1))
print(column_index_from_string("C"))

# Find the total number of rows or columns on a worksheet. Returns answer as an int
print(type(sheet.max_row))
print(sheet.max_column)

# Get a pre defined area from the work sheet
print(sheet['A1': 'C3'])

# Use a loop to print out the area
for row in sheet['A1':'C3']:
    for cell in row:
        # .coordinate shows the cells location, well surprisingly .value shows the cells value.
        print(cell.coordinate, cell.value)

# Print out the values of a certain column or row.

for cell in list(sheet.columns)[1]:
    print(cell.value)

for cell in list(sheet.rows)[1]:
    print(cell.value)


############################################# Editing  Excel #############################################


# Create an empty workbook.
wb_2 = openpyxl.Workbook()

# Select the currently active sheet
sheet = wb_2.active

# Give the sheet a title

sheet.title = "The worlds greatest worksheet!"

# Save the worksheet

wb_2.save('example_copy.xlsx')

# Create or remove sheets

wb_2.create_sheet()

wb_2.create_sheet(index=0, title= "The title")

wb_2.remove_sheet(wb.get_sheet_by_name("sheet to delete here"))

# Writing to cells

sheet['A1'] = "What you want to write"

# Adjust font by creating font object then passing it to .font

from openpyxl.styles import Font

font_obj = Font(name="Times New Roman", bold=True, size=25, italic=True)
sheet['A1'].font = font_obj

# Adjust the size of columns or rows.

sheet.row_dimensions[1].height = 70

sheet.column_dimensions['B'].width = 20

# Merge an area of cells into one. All cells between first and second spot become one.

sheet.merge_cells('A1:D3')
